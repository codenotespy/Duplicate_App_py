import sqlite3
import pandas as pd
import os
import xlsxwriter

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Color


# cursor.execute("UPDATE myapp_maintable SET BASBUGFIYATI = REPLACE(BASBUGFIYATI, ',', '.')")

connection = sqlite3.connect('db.sqlite3')
cursor = connection.cursor()


cursor.execute("DROP TABLE IF EXISTS INPUT")

#To insert in database from excel:
df=pd.read_excel('input.xlsx')
df.to_sql('INPUT',connection, index=False, if_exists="replace")
connection.commit()



cursor.execute("DROP TABLE IF EXISTS OUTPUT")
cursor.execute("CREATE TABLE OUTPUT AS SELECT Malzeme, [Malzeme Açıklaması], [Malzeme Açıklaması(ENG)], Menşei, Fiyat, SUM(Miktar) AS Miktar, GTIP FROM INPUT GROUP BY Malzeme || INPUT.Fiyat")
connection.commit()
cursor.execute("ALTER TABLE OUTPUT ADD Brüt FLOAT")
cursor.execute("UPDATE OUTPUT SET Brüt= Miktar*Fiyat")
cursor.execute("UPDATE OUTPUT SET Fiyat = REPLACE(Fiyat, '.', ',')")
cursor.execute("UPDATE OUTPUT SET Brüt = REPLACE(Brüt, '.', ',')")
connection.commit()




### TO EXPORT FROM SQLITE TO EXCEL:
cursor.execute("SELECT Malzeme, [Malzeme Açıklaması], [Malzeme Açıklaması(ENG)], Menşei, Miktar, Fiyat, Brüt, GTIP FROM OUTPUT")
# TO fetch SELECTED DATA IN THE DATABASE
rows = cursor.fetchall()
# TO CREATE EXCEL FILE
workbook = xlsxwriter.Workbook('output.xlsx')
worksheet = workbook.add_worksheet()
# TO WRITE IN THE CREATED EXCEL
worksheet.write('A1', 'Malzeme')
worksheet.write('B1', 'Malzeme Açıklaması')
worksheet.write('C1', 'Malzeme Açıklaması(ENG)')
worksheet.write('D1', 'Menşei')
worksheet.write('E1', 'Miktar')
worksheet.write('F1', 'Fiyat')
worksheet.write('G1', 'Brüt')
worksheet.write('H1', 'GTIP')
row = 1
col = 0
for module in rows:
    worksheet.write_row(row, col, module)
    row += 1


workbook.close()
connection.close()




# FOR ADDING STYLE
#wb = openpyxl.load_workbook("output.xlsx")
wb = load_workbook("output.xlsx")
#Sheet = wb['Sheet1']
Sheet = wb.active # This is selecting the active sheet in the excel automaticaly.
row_count = Sheet.max_row
column_count = Sheet.max_column
print(str(row_count))
print(str(column_count))



#TO ADD STYLE
border_type = Side(border_style='thin', color="000000")
border = Border(top= border_type, bottom= border_type, right= border_type, left = border_type)
#Sheet['B6'].border = border
fill_pattern = PatternFill(patternType='solid', fgColor='E0FFFF')
font_style = Font(bold=True)

# TO COPY FROM "output.xlsx" FILE TO 'U-List' sheet in outputxd.xlsx
#for r in range(1, 10000):
for r in range(1, int(row_count)+1):
    #for c in range(1, 8):
    for c in range(1, int(column_count)+1):
        Sheet.cell(row=r,column=c).border = border # FOR BORDER STYLE
        Sheet.cell(row=1,column=c).fill = fill_pattern
        Sheet.cell(row=1,column=c).font = font_style
        # I wrote 'r+1' to start from second row instead of first row

#To freeze the first colum:
Sheet.freeze_panes = Sheet['A2']

# TO COMMIT & SAVE THE FILE
wb.save("output.xlsx")





'''
### PART 2= TO COPY & PASTE FROM THE FILE TO outputxd.xlsx
# Before Copy & Paste, We should clear values from previous copy paste..
#wb = openpyxl.load_workbook("outputxd.xlsx")
wb = load_workbook("outputxd.xlsx")
#Sheet1 = wb['Sheet1']
Sheet1 = wb.active # This is selecting the active sheet in the excel automaticaly.
row_count1 = Sheet1.max_row
column_count1 = Sheet1.max_column

# To clear Specific Range of Values in sheet named 'list'
#for ax in Sheet1['A2:G10000']:
for ax in Sheet1['A2:G' + str(row_count1)]:
    for cell in ax:
        cell.value = None
# TO COMMIT & SAVE THE CLEARED FILE
wb.save("outputxd.xlsx")


# File-2 to be paste in:
wb2 = openpyxl.load_workbook("output.xlsx")
Sheet2 = wb2['Sheet1']
row_count = Sheet2.max_row
column_count = Sheet2.max_column
print(str(row_count))
print(str(column_count))



#TO ADD STYLE
border_type = Side(border_style='thin', color="000000")
border = Border(top= border_type, bottom= border_type, right= border_type, left = border_type)
#Sheet1['B6'].border = border

# TO COPY FROM "output.xlsx" FILE TO 'U-List' sheet in outputxd.xlsx
#for r in range(1, 10000):
for r in range(1, int(row_count)):
    #for c in range(1, 8):
    for c in range(1, int(column_count)+1):
        Sheet1.cell(row=r+1,column=c).value = Sheet2.cell(row=r+1,column=c).value
        Sheet1.cell(row=r+1,column=c).border = border # FOR BORDER STYLE
        # I wrote 'r+1' to start from second row instead of first row

# TO COMMIT & SAVE THE FILE
wb.save("outputxd.xlsx")
'''

'''
You can use any of the following values for the border style. hair, dashed, medium, slantDashDot, thick, dotted, dashDotDo, dashDot, double, mediumDashDot, mediumDashed, thin, mediumDashDotDot
https://www.w3schools.com/colors/colors_names.asp

You can use any of the following values for pattern type:
solid, lightGray, lightGrid, darkGrid, darkUp, darkVertical, lightTrellis, lightVertical, gray125, lightHorizontal, mediumGray, darkHorizontal, gray0625, darkGray, lightUp, darkDown, lightDown, darkTrellis

# TO OPEN THE SAVED EXCEL FILE
os.system("start EXCEL.EXE outputxd.xlsx")
'''






















# TO OPEN THE SAVED EXCEL FILE
os.system("start EXCEL.EXE output.xlsx")




